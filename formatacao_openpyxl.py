from openpyxl import load_workbook

# Função para converter valores para float
def parse_value(val):
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        val = val.replace('R$', '').replace(',', '.').strip()
        try:
            return float(val)
        except ValueError:
            print(f"Valor não numérico ignorado: {val}")
            return 0
    return 0

# Carregar a planilha
wb = load_workbook('gastos.xlsx')
ws = wb.active

# Ler os dados das colunas A (Gasto), B (Descrição), C (Valor), D (Tipo de gasto), E (Total gasto)
data = []
for row in ws.iter_rows(min_row=2, max_row=21, values_only=True):  # Limitar a 20 linhas + 1
    # Aceitar linhas se Descrição (índice 1) e Valor (índice 2) não forem nulos
    if len(row) >= 5 and row[1] is not None and row[2] is not None:
        data.append([
            row[0] if len(row) > 0 else '',  # Gasto
            row[1],                          # Descrição
            row[2],                          # Valor
            row[3] if len(row) > 3 else '',  # Tipo de gasto
            row[4] if len(row) > 4 else ''   # Total gasto
        ])
    else:
        print(f"Linha ignorada (células insuficientes ou nulas): {row}")

# Depuração: mostrar todas as linhas lidas
print(f"Total de linhas lidas: {len(data)}")
print("Dados lidos:", data)

# Eliminar duplicatas
conferido = set()
dados = []
for row in data:
    row_tuple = tuple(row)
    if row_tuple not in conferido:
        conferido.add(row_tuple)
        dados.append(row)

# Depuração: mostrar dados após remover duplicatas
print(f"Total de linhas após remover duplicatas: {len(dados)}")
print("Dados sem duplicatas:", dados)

# Somar todos os valores da coluna C (Valor, índice 2)
total_gastos = sum(parse_value(row[2]) for row in dados)

# Adicionar o total ao final da lista
dados.append(['', 'Total', total_gastos, '', ''])

# Somar valores por tipo de descrição (coluna B, índice 1)
total_alimentacao = sum(parse_value(row[2]) for row in dados if row[1] == 'Alimentação')
total_lazer = sum(parse_value(row[2]) for row in dados if row[1] == 'Lazer')
total_transporte = sum(parse_value(row[2]) for row in dados if row[1] == 'Transporte')

# Depuração: mostrar os totais calculados
print(f"Total geral: {total_gastos}")
print(f"Total Alimentação: {total_alimentacao}")
print(f"Total Lazer: {total_lazer}")
print(f"Total Transporte: {total_transporte}")

# Limpar a planilha existente (a partir da linha 2)
ws.delete_rows(2, ws.max_row)

# Escrever os dados sem duplicatas
for i, row in enumerate(dados, start=2):
    ws[f'A{i}'] = row[0]
    ws[f'B{i}'] = row[1]
    ws[f'C{i}'] = row[2]
    ws[f'D{i}'] = row[3]
    ws[f'E{i}'] = row[4]

# Escrever os totais por tipo de descrição na coluna G
if len(dados) >= 4:
    ws['G3'] = total_alimentacao
    ws['G4'] = total_lazer
    ws['G5'] = total_transporte
else:
    print("Aviso: Não há linhas suficientes para escrever os totais nas linhas 3, 4 e 5.")

# Salvar a planilha atualizada
wb.save('gastos_atualizados_2.xlsx')